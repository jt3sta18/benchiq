#!/usr/bin/env python3
"""
Regenerate BenchIQ data modules from an updated spreadsheet export.

Usage:
    python3 scripts/refresh_data.py path/to/updated.xlsx

Reads the lab workbook (location-grid inventory tabs + ordering-log tabs),
normalizes them, adds illustrative CAS + GHS hazard enrichment for well-known
chemicals, and writes:  data/inventory.js, data/orders.js, data/agg.js

Requires: openpyxl   (pip install openpyxl)
"""
import sys, os, re, json
from collections import Counter, defaultdict

try:
    import openpyxl
except ImportError:
    sys.exit("Missing dependency. Run: pip install openpyxl")

HERE = os.path.dirname(os.path.abspath(__file__))
DATA = os.path.join(HERE, "..", "data")

# Which tabs are location-grid inventories, and the storage environment each maps to.
INV_TABS = {
    "Chemical Inventory": "Room temp",
    "4C Inventory": "4°C fridge",
    "-20C Inventory": "-20°C freezer",
}
ENVCODE = {"Room temp": "RT", "4°C fridge": "4C", "-20°C freezer": "-20C"}
ORDER_TABS = [
    "Jan 24- Jun 24", "July 23-Dec23", "Jan 23 - June 23",
    "Jan 22- Jun 22", "Oct 20-Jan 21", "Feb 21 - June 21 ",
]
HEADER_RE = re.compile(r"^(box|shelf|cabinet|rack|inhibitors|contents|drawer|misc|miscellaneous)", re.I)
JUNK_RE = re.compile(r"(pat'?s shit|randomly labeled|dea bag|xxx|^misc(ellaneous)?$)", re.I)

# substring(lower) -> (GHS class, CAS)  — illustrative; verify against the real SDS in production.
HAZ = {
    'sodium hydroxide': ('Corrosive', '1310-73-2'), 'potassium hydroxide': ('Corrosive', '1310-58-3'),
    'ethidium bromide': ('Mutagen / Toxic', '1239-45-8'), 'formaldehyde': ('Carcinogen / Toxic', '50-00-0'),
    'paraformaldehyde': ('Carcinogen / Toxic', '30525-89-4'), 'chloroform': ('Toxic / Carcinogen', '67-66-3'),
    'acrylamide': ('Carcinogen / Toxic', '79-06-1'), 'bisacrylamide': ('Toxic', '110-26-9'),
    'hydrogen peroxide': ('Oxidizer / Corrosive', '7722-84-1'), 'sodium azide': ('Acute Toxic', '26628-22-8'),
    'phenol': ('Toxic / Corrosive', '108-95-2'), 'methanol': ('Flammable / Toxic', '67-56-1'),
    'ethanol': ('Flammable', '64-17-5'), 'acidic ethanol': ('Flammable', '64-17-5'),
    'dmso': ('Irritant', '67-68-5'), 'glutaraldehyde': ('Toxic / Sensitizer', '111-30-8'),
    'sodium dodecyl sulfate': ('Flammable / Irritant', '151-21-3'), 'trypan': ('Suspected Carcinogen', '72-57-1'),
    'crystal violet': ('Toxic / Mutagen', '548-62-9'), 'sodium fluoride': ('Acute Toxic', '7681-49-4'),
    'sodium nitrite': ('Oxidizer / Toxic', '7632-00-0'), 'sodium bisulfite': ('Irritant', '7631-90-5'),
    'puromycin': ('Acute Toxic', '58-58-2'), 'blasticidin': ('Acute Toxic', '3513-03-9'),
    'cycloheximide': ('Toxic / Mutagen', '66-81-9'), 'cyclohexamide': ('Toxic / Mutagen', '66-81-9'),
    'sodium borate': ('Reproductive Toxin', '1330-43-4'), 'sodium tetraborate': ('Reproductive Toxin', '1330-43-4'),
    'imidazole': ('Corrosive / Repro Toxin', '288-32-4'), 'bisphenol a': ('Reproductive Toxin', '80-05-7'),
    'thapsigargin': ('Acute Toxic', '67526-95-8'), 'ammonium peroxydisulfate': ('Oxidizer / Sensitizer', '7727-54-0'),
    'ammonium persulfate': ('Oxidizer / Sensitizer', '7727-54-0'), 'salicylic acid': ('Irritant', '69-72-7'),
    'tetracycline': ('Irritant', '60-54-8'), 'chloramphenicol': ('Suspected Carcinogen', '56-75-7'),
    'guanadine': ('Irritant', ''), 'guanidine': ('Irritant', ''),
}

def clean(s):
    return re.sub(r"\s+", " ", str(s)).strip()

def enrich(name):
    ln = name.lower()
    for k, (cls, cas) in HAZ.items():
        if k in ln:
            return cls, cas
    return "", ""

def canon_vendor(v):
    v = v.lower().strip()
    m = {'fisher': 'Fisher', 'thermofischer': 'Thermo Fisher', 'thermofisher': 'Thermo Fisher',
         'sigma': 'Sigma-Aldrich', 'abcam': 'Abcam', 'nicoya': 'Nicoya', 'rpeptide': 'rPeptide',
         'innovagen': 'Innovagen', 'cayman': 'Cayman', 'promega': 'Promega', 'biorad': 'Bio-Rad', 'enzo': 'Enzo'}
    for k, val in m.items():
        if k in v:
            return val
    return v.title() if v else 'Unknown'

def canon_person(n):
    n = n.lower().strip()
    if n in ('', 'fisher', 'sigma', 'abcam'):
        return None
    n = re.sub(r'dr\.?\s*|sir', '', n).strip()
    return n.title() if n else None

def main(path):
    wb = openpyxl.load_workbook(path, data_only=True)

    # ---- inventory ----
    items = []
    for tab, env in INV_TABS.items():
        if tab not in wb.sheetnames:
            print(f"  (skip: no tab '{tab}')"); continue
        ws = wb[tab]
        rows = [[clean(v) if v not in (None, '') else '' for v in r] for r in ws.iter_rows(values_only=True)]
        cur = {}
        for r in rows:
            if any(c and HEADER_RE.match(c) for c in r):
                for ci, c in enumerate(r):
                    if c and HEADER_RE.match(c):
                        cur[ci] = c
                continue
            for ci, c in enumerate(r):
                if not c or JUNK_RE.search(c) or len(c) < 2:
                    continue
                cls, cas = enrich(c)
                items.append([c, ENVCODE[env], cur.get(ci, ''), cls, cas])
    # dedupe
    seen, dedup = set(), []
    for it in items:
        key = (it[0].lower(), it[1], it[2])
        if key in seen:
            continue
        seen.add(key); dedup.append(it)
    items = dedup

    # ---- orders ----
    def norm(s):
        return re.sub(r"\s+", " ", str(s)).strip().lower()
    orders = []
    for tab in ORDER_TABS:
        if tab not in wb.sheetnames:
            continue
        ws = wb[tab]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        hdr = [norm(x) for x in rows[0]]
        def col(*names):
            for n in names:
                for i, h in enumerate(hdr):
                    if n in h:
                        return i
            return None
        ci = {'name': col('name'), 'vendor': col('company'), 'cat': col('catalog'),
              'item': col('item name', 'description'), 'qty': col('qty'), 'price': col('unit price')}
        for r in rows[1:]:
            g = lambda k: (r[ci[k]] if ci[k] is not None and ci[k] < len(r) else None)
            item = g('item')
            if item in (None, '') and g('cat') in (None, ''):
                continue
            try: price = float(g('price'))
            except (TypeError, ValueError): price = None
            try: qty = float(g('qty'))
            except (TypeError, ValueError): qty = None
            orders.append({'by': str(g('name') or '').strip(), 'vendor': str(g('vendor') or '').strip(),
                           'cat': str(g('cat') or '').strip(), 'item': str(item or '').strip(),
                           'qty': qty, 'price': price, 'period': tab.strip()})

    # ---- aggregates ----
    money = lambda o: (o['price'] or 0) * (o['qty'] or 1)
    sv, sp = defaultdict(float), defaultdict(float)
    for o in orders:
        sv[canon_vendor(o['vendor'])] += money(o)
        p = canon_person(o['by'])
        if p:
            sp[p] += money(o)
    agg = {
        'inv_total': len(items),
        'by_env': dict(Counter({'RT': 'Room temp', '4C': '4°C fridge', '-20C': '-20°C freezer'}[i[1]] for i in items)),
        'n_hazard': sum(1 for i in items if i[3]),
        'orders_total': len(orders),
        'total_spend': round(sum(money(o) for o in orders)),
        'spend_by_vendor': sorted([[k, round(v)] for k, v in sv.items()], key=lambda x: -x[1])[:8],
        'spend_by_person': sorted([[k, round(v)] for k, v in sp.items()], key=lambda x: -x[1])[:6],
    }

    os.makedirs(DATA, exist_ok=True)
    with open(os.path.join(DATA, "inventory.js"), "w") as f:
        f.write("export default " + json.dumps(items, ensure_ascii=False, separators=(',', ':')) + ";\n")
    with open(os.path.join(DATA, "orders.js"), "w") as f:
        f.write("export default " + json.dumps(orders, ensure_ascii=False) + ";\n")
    with open(os.path.join(DATA, "agg.js"), "w") as f:
        f.write("export default " + json.dumps(agg, ensure_ascii=False) + ";\n")

    print(f"✓ inventory: {len(items)} items ({agg['n_hazard']} hazard-flagged)")
    print(f"✓ orders:    {len(orders)} lines  (~${agg['total_spend']:,} tracked spend)")
    print(f"✓ wrote data/inventory.js, data/orders.js, data/agg.js")

if __name__ == "__main__":
    if len(sys.argv) < 2:
        sys.exit("Usage: python3 scripts/refresh_data.py path/to/updated.xlsx")
    main(sys.argv[1])
